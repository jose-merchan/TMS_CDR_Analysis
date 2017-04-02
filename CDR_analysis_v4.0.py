#!/Users/josemerchan/anaconda3/bin/python

import re, openpyxl, datetime, sys
from openpyxl.utils import get_column_letter
from openpyxl import load_workbook
import matplotlib.pylab as plt
import seaborn as sns
import pandas as pd


def concurrent_calls_xlsx(file_name,re_origin,re_destination,duration):
    """
    This function will take a CDR report from TMS and provide the number of concurrent calls.
    User will have to define the columns where the start time and duration are placed within the spreadsheet
    :param filename: Excel file coming from TMS in xlsx format. Note that by default TMS provides the Excel in xls format
    :param re_origin: RegEx to match the origin of the calls
    :param re_destination: RegEx to match the destination of the calls
    :param duration: string representing the minimum call duration in seconds (it must be an integer)
    :return: Create another tab within the spreadsheet showing the results and a graphic (concurrent calls/time)
    :return: Plot an image of the number of concurrent calls matching the search criteria
    """
    try:
        duration = int (duration)
    except:
        print ("Duration should be expressed as string but taking int format")

    #Ask user to identify the columns where Start Time, Duration, Caller and Called party are defined
    start_time_header = input("Type the column where Start Time is defined (A,B,C,...): ").upper()
    duration_header = input("Type the column where Duration in seconds is defined (A,B,C,...): ").upper()
    caller_header = input("Type the column where Calling Party is defined (A,B,C,...): ").upper()
    called_header = input("Type the column where Called Party is defined (A,B,C,...): ").upper()


    try:
        # Load Workbook
        wb = load_workbook(filename= file_name)
        # Load Worksheet
        sheet = wb.get_sheet_by_name(wb.get_sheet_names()[0])
    except (FileNotFoundError, UnboundLocalError) as e:
        print ("No such file or directory: " + e + " " + file_name)

    # Take the rows (filter) to another tab whose origin and destination do match the regex re_origin and re_destination
    column1 = [] # Start time
    column2 = [] # Duration (seconds)
    column3 = [] # Source Number
    column4 = [] # Destination
    #######################
    start_column = [] # Start time
    duration_column = [] # Duration (seconds)
    source_column = [] # Source Number
    destination_column = [] # Destination

    # Create list if the origin matches
    for enum,row in enumerate(sheet['A']):
        if enum == 0:
            continue
        else:
            try:
                compare = re.compile(re_origin, re.IGNORECASE)
                result = compare.search(str(sheet['{0}{1}'.format(caller_header,enum + 1)].value))
            except re.error:
                print ("Review Regex for {}".format(re_origin))
                sys.exit(0)
            if result:
                column1.append(sheet['{0}{1}'.format(start_time_header,enum + 1)].value)
                column2.append(sheet['{0}{1}'.format(duration_header,enum + 1)].value)
                column3.append(sheet['{0}{1}'.format(caller_header,enum + 1)].value)
                if not sheet['{0}{1}'.format(called_header,enum + 1)].value:
                    column4.append("Unknown") # If no value on cell set to Unknown
                else:
                    column4.append(sheet['{0}{1}'.format(called_header,enum + 1)].value)

    # Removing entries from list if the destination does not match and call duration is smaller than duration (seconds)
    for enum, items in enumerate(column4):
        try:
            compare = re.compile(re_destination, re.IGNORECASE)
            result = compare.search(items)
        except re.error:
            print("Review Regex for {}".format(re_destination))
            sys.exit(0)
        if result and (int(column2[enum]) > duration):
            start_column.append(column1[enum])
            duration_column.append(column2[enum])
            source_column.append(column3[enum])
            destination_column.append(column4[enum])

    # Call Analysis time function to calculate number of concurrent calls
    dict_concurrency,origin_time,end_time = analysis(start_column,duration_column,re_origin,re_destination)

    # Creating a new worksheet - Results.
    for names in wb.get_sheet_names():
        if names != 'Results':  # If Result does not exist continue to then create it
            continue
        else:  # If Results exict then remove it from the workbook
            sheet = wb.get_sheet_by_name('Results')
            wb.remove_sheet(sheet)

    wb.create_sheet(title='Results')
    sheet = wb.get_sheet_by_name('Results')
    sheet['A1'] = "Time"
    sheet['B1'] = "Concurrent Calls"

    for enum, keys in enumerate(sorted(dict_concurrency)):
        sheet['A{}'.format(enum + 2)] = keys.strftime(
            "%m-%d-%Y %H:%M:%S %p")  # Converting datetime objects back to string on US time format
        sheet['B{}'.format(enum + 2)] = dict_concurrency[keys]

    # Save the workbook
    wb.save(file_name)

    #Call the printer function to print the result with matplotlib
    printer(origin_time,end_time,dict_concurrency,re_origin,re_destination)

def analysis(start_time,duration,re_origin,re_destination):
    """
    Function that give a list with start and end time (list compound by strings) calculate the number of concurrent calls
    :param start_time: List with the start time that matches the search criteria
    :param duration: List with the duration of the call that matches the search criteria
    :return:
    """
    # If there is a match of the search criteria then the list of start_time_column will not be "None"
    if len(start_time):
        # Taking Call start time
        try:
            time_start = [datetime.datetime.strptime(x, '%m/%d/%Y %H:%M:%S %p') for x in start_time]  # Time in US format month/day/Year - hour in am/pm
        except:
            time_start = [datetime.datetime.strptime(x, '%d/%m/%Y %H:%M:%S') for x in start_time]  # Time in EU format day/month/Year - hour in 24 hours format

        # Taking the duration in seconds
        duration = [datetime.timedelta(seconds=int(x)) for x in duration] # Coverting duration to timedelta. Timedelta takes integer and not string so converting to integer first
        time_start, duration = zip(*sorted(zip(time_start, duration)))  # Make sure the CDRs are listed on the right order
        time_end = [x+y for (x,y) in zip(time_start,duration)] # time_end is equal to time_start + duration

        origin_time = time_start[0] # First sample of our analysis
        end_time = time_end[-1] # Last sample of our analysis

        """
        Organizing data in a dictionary whose keys are the start and end time of calls.
        Values will be the number of calls starting or ending at that time
        """
        dict_list = {}

        for start in time_start:
            if start in dict_list:
                dict_list[start] += 1
            else:
                dict_list[start] = 1

        for end in time_end:
            if end in dict_list:
                dict_list[end] -= 1
            else:
                dict_list[end] = -1

        # Creating other dictionary that calculate the number of concurrent calls using as keys the previous start and end times
        total = 0
        dict_concurrency = {}
        for keys in sorted(dict_list):
            total += dict_list[keys]
            dict_concurrency[keys] = total
        return [dict_concurrency,origin_time,end_time]

    else:
        print("No matches with origin: {}".format(re_origin) + " and/or destination: {}".format(re_destination))
        sys.exit(0)

def printer(origin_time,end_time,dict_concurrency,re_origin,re_destination):
    """
    This function will take care of printing the image based on the origin and end time of the samples
    that match the search criteria in terms of origin, destination and duration. The steps to follow are
    # Creating a dictionary with periodic sample within the time frame analyzed
    # Using to that end the first and last sample of CDR
    :param origin_time: First temporal sample
    :param end_time: Last temporal sample
    :param dict_concurrency: dict containing the number of concurrent calls
    :param re_origin: Regular expression used to match the origin
    :param re_destination: Regular expression used to match the destination of the call
    :return:
    """

    time_frame = end_time - origin_time  # Last sample - First sample give us the interval we are analyzing
    slots = int(time_frame / datetime.timedelta(minutes=5))

    # serie of periodical samples for period comprised between the start and end time with frequency 5 minutes
    date_list = {}
    for i in range(slots):
        stamps = origin_time + i * datetime.timedelta(minutes=5)  # Sampling every 5 minutes
        date_list[stamps] = 0  # Setting the number of concurrent calls to 0

    """
    Introducing the from concurrency to out previously created dictionary that covers the whole period of time to analyze
    If key not in dictionary take previous value
    """
    for i in dict_concurrency:
        date_list[i] = dict_concurrency[i]

    a = 0
    for i in sorted(date_list):
        try:
            a = dict_concurrency[i]
        except KeyError:
            date_list[i] = a

    # Print the result using matplotlib
    fig = plt.figure()
    fig.suptitle(
        'Concurrent calls \n Origin: {}'.format(re_origin) + " " + "and Destination: {}".format(re_destination),
        fontsize=15)
    plt.xlabel('Time', fontsize=22)
    plt.ylabel('Concurrent Calls', fontsize=22)
    x, y = zip(*sorted(date_list.items()))
    sns.set_style("darkgrid")
    plt.plot(x, y, sns.xkcd_rgb["medium green"], lw=2)
    plt.show()

def pandas_concurrent_calls (file_name,re_origin,re_destination,duration):
    """

    :param file_name: FileName where CDRs are contained
    :param re_origin: regex used to match the caller
    :param re_destination: regex used to match destination
    :param duration: minimum call duration
    :return: Excel File with Concurrent Calls and Graphic
    """

    start_time = "Time"
    caller = "Source Number"
    callee = "Destination Number"
    call_length = "Duration (sec)"

    # Columns from the Excel Sheet where the data relative to analysis is contained
    column_names = [start_time, caller, callee, call_length]

    df = pd.read_excel(file_name)
    # Filter using duration
    duration = df[call_length] > int(duration)
    # Filter using regex origin
    regex_origin = df[caller].str.contains(re_origin,
                                                    flags=re.IGNORECASE, regex=True, na=False)
    # Filter using regex destination
    regex_destination = df[callee].str.contains(re_destination,
                                                              flags=re.IGNORECASE, regex=True, na=False)
    # Apply filters to get the calls that match our search criteria
    filtered_df = (df[regex_origin & regex_destination & duration])
    # Drop Columns that do not apply to our analysis
    filtered_df.drop([x for x in filtered_df.columns if x not in column_names],
                     axis=1, inplace=True)

    # Move time from string to datetime
    filtered_df.loc[:, start_time] = pd.to_datetime(filtered_df[start_time], errors='raise')
    # Move duration from string to timedelta
    filtered_df.loc[:, call_length] = pd.to_timedelta(filtered_df[call_length], unit="s", errors='coerce')
    # Sort DataSet by Call Start time
    filtered_df.sort_values(by=[start_time])
    # Determine the time when the call ends
    filtered_df["End Time"] = filtered_df[start_time] + filtered_df[call_length]

    first_call_start_time = filtered_df[start_time].iloc[0]
    last_call_end_time = filtered_df['End Time'].iloc[-1]


    # Every time we have a call starting we create a flag set to 1
    start_time_counter = pd.Series(1, filtered_df[start_time], name="Start Time Counter")
    # Every time a call end we set a flag set to -1
    end_time_counter = pd.Series(-1, filtered_df["End Time"], name="End Time Counter")

    # Create a DataFrame concatenating two series
    filtered_data = pd.concat([start_time_counter], axis=1)
    filtered_data = filtered_data.join(end_time_counter, how='outer')

    # Cumulative sum of the values of two columns. NaN values set to 0
    filtered_data['Concurrent Calls'] = (filtered_data['Start Time Counter'].fillna(0) + \
                                         filtered_data['End Time Counter'].fillna(0)).cumsum()

    # Save Results to Excel file
    if sys.argv[1][-3:].lower()=="xls":
        with pd.ExcelWriter("{}".format(sys.argv[1][:-4]) + '_Concurrent_Calls.xlsx') as writer:
            filtered_data.to_excel(writer, sheet_name="Concurrent Calls")
    elif sys.argv[1][-4:].lower()=="xlsx":
        with pd.ExcelWriter("{}".format(sys.argv[1][:-5]) + '_Concurrent_Calls.xlsx') as writer:
            filtered_data.to_excel(writer, sheet_name="Concurrent Calls")

    # Create new Index from time period comprised between the first and last sample with frequency 1 second
    idx = pd.date_range(filtered_data.index[0], filtered_data.index[-1], freq='S')
    new_series = filtered_data['Concurrent Calls']
    # Remove Duplicate Index from the series and take the last, which will have the right value
    new_series = new_series[~new_series.index.duplicated(keep='last')]
    # Reindex series with new Index value
    new_series.index = pd.DatetimeIndex(new_series.index)
    new_series = new_series.reindex(idx,method='pad')
    #new_series.fillna(method='ffill')

    # Plot graphic
    ax = new_series.plot(title="Concurrent Calls with origin {}".format(re_origin) + " and destination {}".format(re_destination))
    ax.set_xlabel("Time")
    ax.set_ylabel("No of Concurrent Calls")

    #filtered_data.plot(ax=ax)

    plt.show()



if __name__ == "__main__":
    # Verify if XLSX or XLS
    if (sys.argv[1][-4:].lower()=="xlsx" or sys.argv[1][-3:].lower()=="xls"):
        pandas_concurrent_calls(sys.argv[1],r'.*',r'.*', "10000")

    # Verify if CSV
    elif (sys.argv[1][-3:].lower()=="csv"):
        print ("do something")